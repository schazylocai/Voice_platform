import numpy as np
import openpyxl
import streamlit as st
import os
from dotenv import load_dotenv
import stripe
import pandas as pd
import altair as alt
import warnings
from datetime import datetime
import smtplib

from langchain.chat_models import ChatOpenAI
from langchain.memory import ConversationBufferMemory
from langchain.prompts import PromptTemplate
from langchain_experimental.agents.agent_toolkits import create_pandas_dataframe_agent
from langchain.agents.agent_types import AgentType

from src.Change_Text_Style import change_text_style_english

load_dotenv('.env')  # read local .env file
secret_key = os.environ['OPENAI_API_KEY']

stripe_publishable_key = os.environ['STRIPE_PUBLISHABLE_KEY']
strip_secret_key = os.environ['STRIPE_SECRET_KEY']
stripe_api_key = os.environ['STRIPE_API_KEY']
stripe.api_key = strip_secret_key

max_files = 5
final_result = {"question": "", "answer": ""}
violet = "rgb(169, 131, 247)"
red = "rgb(232,89,83)"


def launch_excel_app_eng():
    ######################################### Catch exceptions #########################################
    def catch_exception(file_name):
        with col2:
            st.subheader(f":red[File {file_name} couldn't be loaded. The file has some irregularities!]")
        return False

    ######################################### Clear all files #########################################
    # def clear_all_files():
    #     st.empty()
    #     st.session_state.gpt_excel_messages_eng = []
    #     st.session_state.gpt_excel_chat_history_eng = []
    #     st.session_state.gpt_excel_sheets_frame_eng = pd.DataFrame()
    #     st.session_state.gpt_excel_sheets_frame_eng_adjusted = pd.DataFrame()
    #     st.session_state.gpt_excel_continue_analysis_eng = False
    #     st.session_state.gpt_excel_header_row_index = 0
    #     st.session_state.gpt_excel_file_in_memory = 0
    #     st.session_state.gpt_excel_file_to_upload = None
    #     st.session_state.gpt_excel_output_dataframe = None
    #     st.session_state.gpt_excel_output_chart = None

    #################################### Convert uploaded files to text ####################################
    def convert_excel_to_dataframe(my_file, sheet_idx):
        # Check if the upload file is an Excel file
        try:
            if str(my_file.name).endswith('.xlsx'):
                excel_file_uploaded = openpyxl.load_workbook(my_file, read_only=True)

                warnings.simplefilter("ignore")

                # Iterate through all sheets in the Excel file
                with (st.spinner('Loading in_data...')):
                    sheet_name = excel_file_uploaded.sheetnames[sheet_idx]
                    current_sheet_data = pd.read_excel(
                        my_file,
                        engine='openpyxl',
                        sheet_name=sheet_name,
                        parse_dates=True,
                        keep_default_na=False,
                        dtype_backend='numpy_nullable',
                        nrows=5000,
                    )
                    date_written = []
                    adjusted_frame = current_sheet_data.replace(to_replace=r'^\s*$', value=np.nan, regex=True)
                    adjusted_frame = adjusted_frame.dropna(thresh=1)
                    try:
                        for col in adjusted_frame.columns:
                            new_col_name = col.replace("'", " ")
                            adjusted_frame = adjusted_frame.rename(columns={col: new_col_name})
                    except Exception as e:
                        adjusted_frame = adjusted_frame.dropna(thresh=1)

                    # check if columns are missing some values
                    def is_integer(val):
                        try:
                            float_val = float(val)
                            return float_val.is_integer()
                        except ValueError:
                            return False

                    for col in adjusted_frame.columns:
                        # Calculate the percentage of integer values in the column
                        perc_integers = adjusted_frame[col].apply(
                            lambda x: pd.to_numeric(x, errors='coerce') and pd.to_numeric(x,
                                                                                          errors='coerce') % 1 == 0).mean()

                        # If more than 90% are integers, replace non-integers with 0
                        if perc_integers > 0.9:
                            # Convert the column to numeric, coercing errors to NaN, then fill NaNs with 0
                            adjusted_frame[col] = pd.to_numeric(adjusted_frame[col], errors='coerce').fillna(0)

                            # If you need the column as integers
                            adjusted_frame[col] = adjusted_frame[col].astype(float)

                    # Write all column names in small caps
                    try:
                        adjusted_frame.columns = [frame.strip().lower() for frame in adjusted_frame.columns]
                    except Exception as e:
                        adjusted_frame.columns = adjusted_frame.columns

                    # Check if there is a date column
                    date_cols = adjusted_frame.drop(adjusted_frame.select_dtypes(exclude=[np.datetime64,
                                                                                          datetime]), axis=1)
                    if len(adjusted_frame.columns) > 1:
                        try:
                            date_terms = ['date', 'year', 'month', 'day', 'hour', 'minutes', 'seconds', 'time',
                                          'min', 'sec']
                            date_written = [col for col in adjusted_frame.columns if
                                            any(date_term in col for date_term in date_terms)]

                            if date_cols.empty:
                                final_columns = date_written
                            else:
                                final_columns = date_cols.columns.tolist().append(date_written)
                            for col in [final_columns]:
                                adjusted_frame[col] = adjusted_frame[col].ffill()
                                adjusted_frame[col] = adjusted_frame[col].apply(pd.to_datetime)

                        except Exception as e:
                            st.write()

                    # Check if there is a numerical column
                    non_date_cols = adjusted_frame.drop(adjusted_frame.select_dtypes(include=[np.datetime64,
                                                                                              datetime]), axis=1)
                    numerical_cols_to_add = []
                    for col in non_date_cols.columns:
                        try:
                            adjusted_frame[col] = adjusted_frame[col].fillna(0)
                            adjusted_frame[col] = adjusted_frame[col].apply(pd.to_numeric)
                            numerical_cols_to_add.append(col)
                        except Exception as e:
                            st.write()

                    numerical_cols = adjusted_frame.loc[:, numerical_cols_to_add]

                    # replace null values in categorical column with 'NA'
                    categorical_cols = adjusted_frame.drop(date_cols, axis=1)
                    categorical_cols = adjusted_frame.drop(numerical_cols, axis=1)
                    categorical_cols = categorical_cols.fillna('Missing Data')

                    for col in categorical_cols.columns:
                        adjusted_frame[col] = categorical_cols[col]

                    adjusted_frame = adjusted_frame.reset_index(drop=True)
                    # convert an Excel sheet to pdf

                    return adjusted_frame

            else:
                catch_exception(my_file.name)
                return []

        except Exception as e:
            catch_exception(my_file.name)
            st.write(e)
            return []

    ######################################### Compose layout #########################################
    col1, col2, col3 = st.columns(3)

    with col1:
        st.title(":red[GPT Excel Analyzer]")

    st.subheader(":violet[Ensure that the :red[first row] of the table has the :red[column names], "
                 "and that there are :red[no empty rows] at the end of the table.]")

    with col2:
        ################################# load documents #################################
        # upload file 1
        try:
            excel_file_1 = st.sidebar.file_uploader(
                label=':violet[Maximum allowed rows = 5,000]',
                type=['xlsx'],
                accept_multiple_files=False, key='excel_file_1_eng',
                label_visibility='visible')

        except Exception as e:
            gpt_excel_send_email_error(e)

        retry_count_1 = 0
        if excel_file_1:
            st.session_state.gpt_excel_file_to_upload = excel_file_1
            with st.session_state.gpt_excel_file_to_upload:
                st.session_state.gpt_excel_file_in_memory = 1
                try:
                    if str(excel_file_1.name).endswith('.xlsx'):
                        excel_file = openpyxl.load_workbook(excel_file_1, read_only=True)

                        # Create a list of sheet names and their corresponding indices
                        sheet_info = [(i, sheet_name) for i, sheet_name in enumerate(excel_file.sheetnames)]

                        # Create a radio button to choose a sheet using its index
                        st.sidebar.subheader(':violet[Choose a sheet:]')
                        choose_sheet_index = st.sidebar.selectbox(
                            label='Choose a sheet from the uploaded Excel file:',
                            options=[sheet for sheet in sheet_info],
                            format_func=lambda x: x[1],  # Display sheet names
                            key='choose_sheet_index',
                            label_visibility='hidden',
                        )

                        # Get the selected sheet index
                        selected_index, selected_sheet_name = choose_sheet_index
                        st.session_state.gpt_excel_continue_analysis_eng = True

                        st.session_state.gpt_excel_sheets_frame_eng = convert_excel_to_dataframe(
                            excel_file_1, selected_index)

                except Exception as e:
                    # st.sidebar.write("Maximum retry attempts reached. Upload failed.")
                    gpt_excel_send_email_error(e)

        else:
            st.session_state.gpt_excel_sheets_frame_eng = pd.DataFrame()

    # with col3:
    #     # set the clear button
    #     st.write("")
    #     st.write("")
    #     clear = st.button(':white[Clear conversation & memory]', key='clear', use_container_width=True)
    #
    #     if clear:
    #         clear_all_files()

    ######################################### render tables #########################################
    sheets_frame = st.session_state.gpt_excel_sheets_frame_eng

    if excel_file_1 and st.session_state['gpt_excel_file_in_memory'] == 1:
        graph_type = 'None'
        col_A = 'None'
        col_B = 'None'

        st.write(':violet[You can :red["edit any value"] in the table. You can :red["add rows"] to the table by '
                 'clicking on the :red["＋"] icon at the bottom of the table. You can :red["delete rows"] in the table '
                 'by ticking :red["✓"] on the far left box of each desired row, and then clicking on the :red["trash '
                 'bin"] icon :red["🗑️"] on the far top right icons above the table. You can :red["save the table"] '
                 'by clicking on the Download icon :red["↓"], :red["search for values"] in the table by clicking on '
                 'the :red["Loupe"] icon :red["🔍"], and :red["view the table in full screen"] by clicking on the '
                 'square icon :red["⌑"]: with the icons located at the far top right icons above the table.]')
        st.divider()

        # Validate the dataframe
        def show_dataframe():
            try:
                st.session_state['gpt_excel_sheets_frame_eng'] = st.data_editor(
                    data=st.session_state['gpt_excel_sheets_frame_eng'],
                    use_container_width=True, height=500,
                    num_rows='dynamic', hide_index=False,
                    key='editor').reset_index(drop=True)

            except Exception as e:
                st.session_state['gpt_excel_sheets_frame_eng'] = st.data_editor(
                    data=st.session_state['gpt_excel_sheets_frame_eng'],
                    use_container_width=True, height=500,
                    num_rows='dynamic', hide_index=False,
                    key='editor_except')

        show_dataframe()
        sheets_frame = st.session_state['gpt_excel_sheets_frame_eng']

    if excel_file_1 and st.session_state['gpt_excel_file_in_memory'] == 1:
        # select a graph type
        st.sidebar.subheader(':violet[Choose a graph type:]')
        graph_type = st.sidebar.selectbox(
            label='Choose a graph type',
            options=['None', 'bar chart', 'line chart', 'pie chart'],
            key='choose_type',
            label_visibility='hidden',
        )
        if graph_type == 'None':
            col_A = 'None'
            col_B = 'None'

        elif graph_type == 'bar chart':
            try:
                if len(sheets_frame.columns) > 1:
                    # select 1st column
                    col_A = st.sidebar.selectbox(
                        label='Choose first column (x-axis / distribution)',
                        options=['None'] + [sheet for sheet in
                                            sheets_frame.select_dtypes(exclude=[np.datetime64, datetime])],
                        key='choose_categorical_A_column',
                        label_visibility='visible',
                    )
                    # select second column
                    col_B = st.sidebar.selectbox(
                        label='Choose second column (y-axis / value)',
                        options=['None'] + [sheet for sheet in sheets_frame.select_dtypes(include='number') if
                                            sheet != col_A],
                        key='choose_value_A_column',
                        label_visibility='visible',
                    )
                    # Isolate dataframe with the selected variables
                    st.divider()
                    # output graph
                    if col_B == 'None':
                        graph_dataframe = pd.DataFrame(data=sheets_frame.loc[:, col_A], columns=[col_A])
                        chart = alt.Chart(graph_dataframe, height=800).mark_bar(cornerRadius=10).encode(
                            alt.X(f'{col_A}:N', axis=alt.Axis(
                                labelFontSize=16,
                                grid=False,
                                labelColor='cyan',
                                titleColor='cyan',
                                titleFontSize=16)),
                            alt.Y('count():Q', axis=alt.Axis(
                                labelFontSize=16,
                                grid=False,
                                labelColor='cyan',
                                titleColor='cyan',
                                titleFontSize=16)),
                            color=f'{col_A}:N'
                        )
                        # Display the chart in Streamlit
                        st.altair_chart(chart, use_container_width=True)
                    else:
                        graph_dataframe = pd.DataFrame(data=sheets_frame.loc[:, [col_A, col_B]],
                                                       columns=[col_A, col_B])
                        chart = alt.Chart(graph_dataframe, height=800).mark_bar(cornerRadius=10).encode(
                            alt.X(f'{col_A}:N', axis=alt.Axis(
                                labelFontSize=16,
                                grid=False,
                                labelColor='cyan',
                                titleColor='cyan',
                                titleFontSize=16)),
                            alt.Y(f'{col_B}:Q', axis=alt.Axis(
                                labelFontSize=16,
                                grid=False,
                                labelColor='cyan',
                                titleColor='cyan',
                                titleFontSize=16)),
                            color=f'{col_A}:N'
                        )
                        # Display the chart in Streamlit
                        st.altair_chart(chart, use_container_width=True)

                else:
                    st.sidebar.write(
                        ':red[Could not create a bar chart! Check the columns in the excel file. You need to '
                        'have at least 2 columns in the file (a category and a value).]')

            except Exception as e:
                st.write('')

        elif graph_type == 'line chart':
            try:
                if len(sheets_frame.columns) > 1:
                    # select 1st column
                    col_A = st.sidebar.selectbox(
                        label='Choose date column (x-axis)',
                        options=['None'] + [sheet for sheet in
                                            sheets_frame.select_dtypes(include=[np.datetime64, datetime])],
                        key='choose_date_column',
                        label_visibility='visible',
                    )
                    if col_A != 'None':
                        # select second column
                        my_vars = [sheet for sheet in sheets_frame.select_dtypes(include=['number'])]
                        col_B = st.sidebar.multiselect(
                            label='Choose value over time column',
                            options=set(my_vars),
                            key='choose_value_B_column',
                            label_visibility='visible',
                        )
                        if col_B:
                            # Melt the DataFrame to transform columns into rows
                            data_A = pd.DataFrame(data=sheets_frame.loc[:, col_A]).reset_index(drop=True)
                            data_B = pd.DataFrame(data=sheets_frame.loc[:, col_B]).reset_index(drop=True)

                            all_data = pd.concat([data_A, data_B], axis=1)
                            all_data = pd.melt(all_data, id_vars=[col_A], var_name='variables',
                                               value_name='points')
                            st.divider()

                            # output graph
                            chart = alt.Chart(all_data, height=600).mark_line(interpolate='monotone',
                                                                              strokeWidth=3,
                                                                              opacity=0.5).encode(
                                alt.X(f'{col_A}:T', axis=alt.Axis(
                                    labelFontSize=14,
                                    grid=False,
                                    labelColor='cyan',
                                    titleColor='cyan',
                                    titleFontSize=18)),
                                alt.Y('points:Q', axis=alt.Axis(
                                    labelFontSize=12,
                                    grid=False,
                                    labelColor='cyan',
                                    titleColor='cyan',
                                    titleFontSize=18)),
                                color='variables:N'
                            )
                            # Display the chart in Streamlit
                            st.altair_chart(chart, use_container_width=True)

                        else:
                            st.sidebar.write(
                                ':red[Could not create a line chart! Check the columns in the excel file. You need to '
                                'have at least 2 columns in the file (a date and a value). ]')

                    else:
                        st.sidebar.write(
                            ':red[Could not create a line chart! Check the columns in the excel file. You need to '
                            'have at least 2 columns in the file (a date and a value). ]')

                else:
                    st.sidebar.write(
                        ':red[Could not create a line chart! Check the columns in the excel file. You need to '
                        'have at least 2 columns in the file (a date and a value). ]')

            except Exception as e:
                st.write('')

        elif graph_type == 'pie chart':
            try:
                # select column
                col_A = st.sidebar.selectbox(
                    label='Choose first column (distribution)',
                    options=['None'] + [sheet for sheet in sheets_frame.select_dtypes(exclude=[np.datetime64,
                                                                                               datetime,
                                                                                               'number'])],
                    key='choose_categorical_A_column',
                    label_visibility='visible',
                )
                # Isolate dataframe with the selected variables
                graph_dataframe = pd.DataFrame(data=sheets_frame.loc[:, col_A], columns=[col_A])
                st.divider()

                # Create a new DataFrame with the counts and percentages
                count_df = graph_dataframe[col_A].value_counts(normalize=True).reset_index()
                count_df.columns = ['Variable', 'Percentage']
                count_df['Percentage'] = (count_df['Percentage'] * 100).round(2)

                # output graph
                chart = alt.Chart(count_df, height=700).mark_arc(innerRadius=100, padAngle=0).encode(
                    theta=alt.Theta('Percentage:Q'),
                    color=alt.Color('Variable:N', scale=alt.Scale(scheme='category20c')),
                    tooltip=['Variable', 'Percentage'],
                )
                # Display the chart in Streamlit
                st.altair_chart(chart, use_container_width=True)

            except Exception as e:
                st.write('')

        else:
            graph_type = 'None'
            st.sidebar.write(':red[Choose graph type.]')

    # st.divider()

    ####################################### Write chat history #######################################
    for message in st.session_state.gpt_excel_messages_eng:
        with st.chat_message(message['role']):
            change_text_style_english(message['content'], 'main_text_white', 'white')

    ######################################## documents ########################################
    if st.session_state.gpt_excel_continue_analysis_eng:

        #################################### Templates ####################################

        response_template = """
                        - you are provided with a dataframe {{sheets_frame}}
                        - This dataframe has to be read in a horizontal and vertical sense to understand the context.
                        - The horizontal lines are the rows of the dataframe.
                        - The vertical lines are the columns of the dataframe.
                        - Take a deep breath and work on this problem step-by-step.

                        - You are only allowed to use the dataframe {{sheets_frame}} given to you.
                        - Don't use any information outside the given dataframe {{sheets_frame}}.

                        - Give your final solution in bullet points.
                        - In your solution, sort the in_lines in descending order.


                        <ctx>
                        {context}
                        </ctx>
                        --------
                        <hs>
                        {history}
                        </hs>
                        --------
                        Answer:
                        """

        prompt_files = PromptTemplate(template=response_template,
                                      input_variables=["history", "context"])

        llm = ChatOpenAI(temperature=0.0, model=st.session_state.ChatOpenAI)
        query_model = create_pandas_dataframe_agent(
            llm,
            sheets_frame,
            agent_type=AgentType.OPENAI_FUNCTIONS,
            agent_executor_kwargs={
                "prompt": prompt_files,
                "memory": ConversationBufferMemory(memory_key="history",
                                                   return_messages=True)}
        )

        #################################### Run the model ####################################
        def create_text_question():

            user_input = st.chat_input('Start querying the tables here...',
                                       max_chars=500, key='user_input')
            if user_input:
                with st.chat_message('user'):
                    st.markdown(user_input)

                st.session_state.gpt_excel_messages_eng.append({'role': 'user', 'content': user_input})

                with st.spinner(text=":red[Query submitted. This may take a minute while we query the table...]"):
                    with st.chat_message('assistant'):
                        try:
                            message_placeholder = st.empty()
                            all_results = ''
                            output = query_model(user_input)
                            user_query = user_input
                            result = output['output']
                            st.session_state.gpt_excel_chat_history_eng.append((user_query, result))
                            all_results += result
                            font_link_eng = '<link href="https://fonts.googleapis.com/css2?family=Roboto:wght@400;700&display=swap" rel="stylesheet">'
                            font_family_eng = "'Roboto', sans-serif"
                            message_placeholder.markdown(
                                f"""
                                                                    {font_link_eng}
                                                                    <style>
                                                                        .bold-text {{
                                                                            font-family: {font_family_eng};
                                                                            font-size: 22px;
                                                                            color: 'white;
                                                                            text-align: left;
                                                                            line-height: 2.2;
                                                                            font-weight: 400;
                                                                        }}
                                                                    </style>
                                                                    <div class="bold-text"><bdi>{all_results}</bdi></div>
                                                                    """, unsafe_allow_html=True)

                            st.session_state.gpt_excel_messages_eng.append(
                                {'role': 'assistant', 'content': all_results})
                            # return result

                        except Exception as e:
                            st.write(":red[Couldn't process the request. Please try again!]")
                            st.write(e)

        create_text_question()

    else:
        st.empty()


def gpt_excel_send_email_error(body):
    sender = 'GPT Excel'
    recipient = os.environ["MY_EMAIL_ADDRESS"]
    subject = 'Error message'
    sender_email = 'samuel'
    with smtplib.SMTP("smtp.gmail.com", 587) as server:
        server.ehlo()
        server.starttls()
        server.login(recipient, os.environ["EMAIL_PASSWORD"])
        server.sendmail(sender, recipient, f"Subject: {subject}\n\n{sender}: {sender_email}\n\n{body}")
